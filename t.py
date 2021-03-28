#!/usr/bin/env python3
#-*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
import sys

source = "收文登记表.xlsx"
template = "文件处理单（模板）.xlsx"
target_file = "文件处理单.xlsx"

#source = 'from.xlsx'
#template = 'to.xlsx'
#target_file = None

def row_valid(row: str):
	if row[0] is None:
		return False
	valid_chars = "年月日"
	for i in valid_chars:
		if i not in row[0]:
			return False
	return True

def get_valid_rows(source: str, row_num_list: list = None) -> list:
	wb = load_workbook(source)
	ws = wb.active
	ret = []
	print("get_valid_rows", row_num_list)
	cur_row_num = 1
	for row in ws.iter_rows(values_only=True):
		if row_num_list is not None and cur_row_num in row_num_list:
			ret.append(row)
		cur_row_num += 1

	return ret

def get_row_count(source: str) -> int:
	wb = load_workbook(source)
	ws = wb.active
	ret = []
	cnt = 0
	for row in ws.iter_rows(values_only=True):
		cnt += 1
	return cnt

def gen_from_template(template: str, rows: list, gen_file_name: str = None):
	"""
	if gen_file_name is not None:
		import shutil
		shutil.copyfile(template, gen_file_name)
		write_to = gen_file_name
	else:
		write_to = template
		"""
	wb = load_workbook(template)
	ws = wb.active

	print(template, rows)
	"row: ('2021年3月26日', '20210001', '党委1', 'xx同志批示清样（在xx上的批示）') "
	for r in rows:
		new_ws= wb.copy_worksheet(ws)
		new_ws.title = r[1]

		new_ws['C5'] = r[2] # 来电单位
		new_ws['F5'] = r[4] # 文件编号
		new_ws['C6'] = r[1] # 收文号
		new_ws['E6'] = r[0] # 收文date
		new_ws['K6'] = r[5] # 密级
		new_ws['C7'] = r[3]

	if gen_file_name:
		write_to = gen_file_name
	else:
		write_to = template
	wb.remove_sheet(ws)
	wb.save(write_to)

	return None

def usage():
	print(sys.argv[0], " [-s source file] [-t template file] [-m target_file] ROW_SELECTOR")
	help_str="""
ROW_SELECTOR:
	N: row N
	N-M: row N to M(included)
	N-: row N to the end
examples:
	1. 生成1 3 4 行
		{prog} 1,3,4
	2. 生成 1-15行
		{prog} 1-15
	2. 生成 1 3 以及8-15行
		{prog} 1,3,8-15
""".format(prog=sys.argv[0])
	print(help_str)

if __name__ == "__main__":
	ranges = None
	row_num_list = []
	total_cnt = get_row_count(source)
	if 1 == len(sys.argv):
		usage()
		exit(0)
	for _i in range(len(sys.argv)):
		if _i == 0:
			continue
		i = sys.argv[_i]
		if i.startswith('-'):
			if i.strip() == "-s":
				source = sys.argv[_i+1]
			elif i.strip() == "-t":
				template = sys.argv[_i+1]
			elif i.strip() == "-m":
				target_file = sys.argv[_i+1]
		else:
			ranges = i

	# ranges = 1,29,20,28-
	print(ranges)
	range_list = ranges.split(",")
	for i in range_list:
		if "-" not in i:
			try:
				num = int(i)
			except:
				print("invalid args: ", i)
				exit(0)
			row_num_list.append(num)
		else:
			# for case 100- 101
			if i.endswith("-"):
				i = i.strip("-")
				start = int(i)
				end = total_cnt
			else:
				split_parts = i.split("-")
				start, end = split_parts
				start = int(start)
				end = int(end)

			for j in range(start, end + 1):
				if j not in row_num_list:
					row_num_list.append(j)
	print("row num list:", row_num_list)
	if target_file is None:
		target_file = template

	if len(row_num_list):
		valid_rows = get_valid_rows(source, row_num_list)
	else:
		valid_rows = get_valid_rows(source)
	print("valid rows:", '\n', valid_rows)
	gen_from_template(template, valid_rows, target_file)
